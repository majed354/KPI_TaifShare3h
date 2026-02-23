#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Create comprehensive KPI Excel workbook with dashboard."""

import csv
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart, PieChart
from openpyxl.chart.series import SeriesLabel

# ── Config ──────────────────────────────────────────────────────────────
CSV_PATH = os.path.join('KPI_TaifShare3h-main', 'data', 'data.csv')
GRADS_CSV = os.path.join('KPI_TaifShare3h-main', 'data', 'graduates_detail.csv')
NONCOMP_CSV = os.path.join('KPI_TaifShare3h-main', 'data', 'non_completers.csv')
OUT_PATH = os.path.join('KPI_TaifShare3h-main', 'data', 'KPI_Data_Complete.xlsx')
YEAR_ORDER = [38, 39, 40, 41, 42, 44, 45, 46, 47]
YEAR_LABELS = {
    38: '1438', 39: '1439', 40: '1440', 41: '1441',
    42: '1442', 44: '1444', 45: '1445', 46: '1446', 47: '1447'
}

# ── Styles ──────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill('solid', fgColor='0D6E6E')
HEADER_FONT = Font(name='Tajawal', bold=True, color='FFFFFF', size=12)
TITLE_FONT = Font(name='Tajawal', bold=True, size=16, color='0D6E6E')
SUBTITLE_FONT = Font(name='Tajawal', bold=True, size=13, color='333333')
DATA_FONT = Font(name='Tajawal', size=11)
NUM_FONT = Font(name='Tajawal', size=11)
PERCENT_FONT = Font(name='Tajawal', size=11, bold=True)
GOLD_FILL = PatternFill('solid', fgColor='C9A227')
LIGHT_FILL = PatternFill('solid', fgColor='E8F5F5')
ALT_FILL = PatternFill('solid', fgColor='F5F5F5')
GREEN_FILL = PatternFill('solid', fgColor='D4EDDA')
YELLOW_FILL = PatternFill('solid', fgColor='FFF3CD')
RED_FILL = PatternFill('solid', fgColor='F8D7DA')
BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)
THICK_BORDER = Border(
    left=Side(style='medium', color='0D6E6E'),
    right=Side(style='medium', color='0D6E6E'),
    top=Side(style='medium', color='0D6E6E'),
    bottom=Side(style='medium', color='0D6E6E'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=True)

# ── Read CSV ────────────────────────────────────────────────────────────
def read_data():
    rows = []
    with open(CSV_PATH, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=';')
        for r in reader:
            row = {}
            row['dept'] = r['Dept_aName']
            row['major'] = r['Major_aName']
            row['degree'] = r['Degree_aName']
            row['sem'] = int(r['Semester'])
            for k in ['students_total','students_male','students_female',
                       'students_saudi','students_international',
                       'students_new','students_retained',
                       'graduates_total','graduates_ontime',
                       'prev_new_count','new_4_ago_count']:
                val = r.get(k, '')
                row[k] = int(val) if val not in ('', None) else 0
            rows.append(row)
    return rows

# ── Helper ──────────────────────────────────────────────────────────────
def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

def style_data_cell(ws, row, col, is_alt=False, is_num=False, is_pct=False):
    cell = ws.cell(row=row, column=col)
    cell.font = PERCENT_FONT if is_pct else (NUM_FONT if is_num else DATA_FONT)
    cell.alignment = CENTER if is_num or is_pct else RIGHT
    cell.border = BORDER
    if is_alt:
        cell.fill = ALT_FILL
    if is_pct:
        cell.number_format = '0.0%'

def auto_width(ws, min_w=12, max_w=30):
    for col in ws.columns:
        mx = min_w
        for cell in col:
            if cell.value:
                mx = max(mx, min(len(str(cell.value)) + 4, max_w))
        ws.column_dimensions[get_column_letter(col[0].column)].width = mx

def pct(num, den):
    if den and den > 0:
        return num / den
    return None

# ══════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════
data = read_data()
wb = Workbook()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  SHEET 1: البيانات الخام (Raw Data)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws1 = wb.active
ws1.title = 'البيانات الخام'
ws1.sheet_view.rightToLeft = True

headers_ar = [
    'القسم', 'التخصص', 'الدرجة', 'السنة',
    'إجمالي الطلاب', 'ذكور', 'إناث',
    'سعوديون', 'دوليون',
    'طلاب جدد', 'مستمرون',
    'إجمالي الخريجين', 'خريجون في الوقت',
    'جدد السنة السابقة', 'حجم دفعة التخرج',
    'نسبة التخرج في الوقت', 'نسبة الاستبقاء',
    'نسبة الذكور', 'نسبة الإناث', 'نسبة الدوليين'
]

# Title
ws1.merge_cells('A1:T1')
ws1.cell(1, 1, 'البيانات الخام - مؤشرات الأداء الأكاديمي').font = TITLE_FONT
ws1.cell(1, 1).alignment = CENTER
ws1.row_dimensions[1].height = 35

# Headers row 2
for i, h in enumerate(headers_ar, 1):
    ws1.cell(2, i, h)
style_header_row(ws1, 2, len(headers_ar))

# Data rows
for idx, d in enumerate(data):
    r = idx + 3
    is_alt = idx % 2 == 1
    year_label = YEAR_LABELS.get(d['sem'], str(1400 + d['sem']))

    grad_rate = pct(d['graduates_ontime'], d['new_4_ago_count'])
    retention = pct(d['students_retained'], d['prev_new_count'])
    male_pct = pct(d['students_male'], d['students_total'])
    female_pct = pct(d['students_female'], d['students_total'])
    intl_pct = pct(d['students_international'], d['students_total'])

    vals = [
        d['dept'], d['major'], d['degree'], year_label,
        d['students_total'], d['students_male'], d['students_female'],
        d['students_saudi'], d['students_international'],
        d['students_new'], d['students_retained'],
        d['graduates_total'], d['graduates_ontime'],
        d['prev_new_count'], d['new_4_ago_count'],
        grad_rate, retention,
        male_pct, female_pct, intl_pct
    ]
    for c, v in enumerate(vals, 1):
        ws1.cell(r, c, v)
        is_pct = c >= 16
        is_num = 5 <= c <= 15
        style_data_cell(ws1, r, c, is_alt, is_num, is_pct)

auto_width(ws1)
ws1.auto_filter.ref = f'A2:T{len(data)+2}'

print(f'Sheet 1: {len(data)} rows written')

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  SHEET 2: ملخص حسب السنة (Year Summary)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws2 = wb.create_sheet('ملخص حسب السنة')
ws2.sheet_view.rightToLeft = True

# Aggregate by year
year_agg = {}
for d in data:
    y = d['sem']
    if y not in year_agg:
        year_agg[y] = {
            'total': 0, 'male': 0, 'female': 0,
            'saudi': 0, 'intl': 0, 'new': 0, 'retained': 0,
            'grads': 0, 'grads_ontime': 0,
            'prev_new': 0, 'new_4_ago': 0, 'programs': 0
        }
    a = year_agg[y]
    a['total'] += d['students_total']
    a['male'] += d['students_male']
    a['female'] += d['students_female']
    a['saudi'] += d['students_saudi']
    a['intl'] += d['students_international']
    a['new'] += d['students_new']
    a['retained'] += d['students_retained']
    a['grads'] += d['graduates_total']
    a['grads_ontime'] += d['graduates_ontime']
    a['prev_new'] += d['prev_new_count']
    a['new_4_ago'] += d['new_4_ago_count']
    a['programs'] += 1

# Title
ws2.merge_cells('A1:N1')
ws2.cell(1, 1, 'ملخص مؤشرات الأداء حسب السنة').font = TITLE_FONT
ws2.cell(1, 1).alignment = CENTER
ws2.row_dimensions[1].height = 35

year_headers = [
    'السنة', 'عدد البرامج', 'إجمالي الطلاب',
    'ذكور', 'إناث', 'سعوديون', 'دوليون',
    'طلاب جدد', 'مستمرون',
    'إجمالي الخريجين', 'خريجون في الوقت',
    'نسبة التخرج', 'نسبة الاستبقاء', 'نسبة الدوليين'
]
for i, h in enumerate(year_headers, 1):
    ws2.cell(2, i, h)
style_header_row(ws2, 2, len(year_headers))

row_n = 3
for y in YEAR_ORDER:
    if y not in year_agg:
        continue
    a = year_agg[y]
    is_alt = (row_n - 3) % 2 == 1

    grad_r = pct(a['grads_ontime'], a['new_4_ago'])
    ret_r = pct(a['retained'], a['prev_new'])
    intl_r = pct(a['intl'], a['total'])

    vals = [
        YEAR_LABELS[y], a['programs'], a['total'],
        a['male'], a['female'], a['saudi'], a['intl'],
        a['new'], a['retained'],
        a['grads'], a['grads_ontime'],
        grad_r, ret_r, intl_r
    ]
    for c, v in enumerate(vals, 1):
        ws2.cell(row_n, c, v)
        is_pct = c >= 12
        is_num = 2 <= c <= 11
        style_data_cell(ws2, row_n, c, is_alt, is_num, is_pct)
    row_n += 1

auto_width(ws2)

# ── Chart: students trend by year ──
chart1 = BarChart()
chart1.type = 'col'
chart1.title = 'إجمالي الطلاب حسب السنة'
chart1.y_axis.title = 'عدد الطلاب'
chart1.x_axis.title = 'السنة'
chart1.style = 10
data_ref = Reference(ws2, min_col=3, min_row=2, max_row=row_n-1)
cats_ref = Reference(ws2, min_col=1, min_row=3, max_row=row_n-1)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.shape = 4
chart1.width = 20
chart1.height = 12
ws2.add_chart(chart1, f'A{row_n + 2}')

# ── Chart: graduation & retention rates ──
chart2 = LineChart()
chart2.title = 'نسب التخرج والاستبقاء'
chart2.y_axis.title = 'النسبة'
chart2.x_axis.title = 'السنة'
chart2.style = 10
grad_ref = Reference(ws2, min_col=12, min_row=2, max_row=row_n-1)
ret_ref = Reference(ws2, min_col=13, min_row=2, max_row=row_n-1)
chart2.add_data(grad_ref, titles_from_data=True)
chart2.add_data(ret_ref, titles_from_data=True)
chart2.set_categories(cats_ref)
chart2.width = 20
chart2.height = 12
ws2.add_chart(chart2, f'A{row_n + 18}')

print(f'Sheet 2: {row_n - 3} year rows')

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  SHEET 3: ملخص حسب البرنامج (Program Summary)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws3 = wb.create_sheet('ملخص حسب البرنامج')
ws3.sheet_view.rightToLeft = True

# Get unique programs
programs = {}
for d in data:
    key = (d['dept'], d['major'], d['degree'])
    if key not in programs:
        programs[key] = []
    programs[key].append(d)

ws3.merge_cells('A1:P1')
ws3.cell(1, 1, 'ملخص مؤشرات الأداء حسب البرنامج (أحدث سنة متاحة)').font = TITLE_FONT
ws3.cell(1, 1).alignment = CENTER
ws3.row_dimensions[1].height = 35

prog_headers = [
    'القسم', 'التخصص', 'الدرجة',
    'أحدث سنة', 'إجمالي الطلاب',
    'ذكور', 'إناث', 'سعوديون', 'دوليون',
    'طلاب جدد', 'مستمرون',
    'خريجون', 'خريجون في الوقت',
    'نسبة التخرج', 'نسبة الاستبقاء',
    'عدد السنوات'
]
for i, h in enumerate(prog_headers, 1):
    ws3.cell(2, i, h)
style_header_row(ws3, 2, len(prog_headers))

row_n = 3
for (dept, major, degree), rows in sorted(programs.items()):
    latest = max(rows, key=lambda x: x['sem'])
    is_alt = (row_n - 3) % 2 == 1

    grad_r = pct(latest['graduates_ontime'], latest['new_4_ago_count'])
    ret_r = pct(latest['students_retained'], latest['prev_new_count'])

    vals = [
        dept, major, degree,
        YEAR_LABELS.get(latest['sem'], str(1400 + latest['sem'])),
        latest['students_total'],
        latest['students_male'], latest['students_female'],
        latest['students_saudi'], latest['students_international'],
        latest['students_new'], latest['students_retained'],
        latest['graduates_total'], latest['graduates_ontime'],
        grad_r, ret_r,
        len(rows)
    ]
    for c, v in enumerate(vals, 1):
        ws3.cell(row_n, c, v)
        is_pct = c in (14, 15)
        is_num = 5 <= c <= 13 or c == 16
        style_data_cell(ws3, row_n, c, is_alt, is_num, is_pct)
    row_n += 1

auto_width(ws3)
ws3.auto_filter.ref = f'A2:P{row_n - 1}'

print(f'Sheet 3: {row_n - 3} programs')

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  SHEET 4: لوحة المعلومات (Dashboard)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws4 = wb.create_sheet('لوحة المعلومات')
ws4.sheet_view.rightToLeft = True

# ── Section 1: KPI Summary Cards (Latest Year = 1447) ──
latest_year = 47
latest_data = [d for d in data if d['sem'] == latest_year]

totals = {
    'students': sum(d['students_total'] for d in latest_data),
    'male': sum(d['students_male'] for d in latest_data),
    'female': sum(d['students_female'] for d in latest_data),
    'saudi': sum(d['students_saudi'] for d in latest_data),
    'intl': sum(d['students_international'] for d in latest_data),
    'new': sum(d['students_new'] for d in latest_data),
    'retained': sum(d['students_retained'] for d in latest_data),
    'grads': sum(d['graduates_total'] for d in latest_data),
    'grads_ontime': sum(d['graduates_ontime'] for d in latest_data),
    'prev_new': sum(d['prev_new_count'] for d in latest_data),
    'new_4_ago': sum(d['new_4_ago_count'] for d in latest_data),
    'programs': len(latest_data),
}

ws4.merge_cells('A1:H1')
ws4.cell(1, 1, f'لوحة المعلومات - مؤشرات الأداء الرئيسية {YEAR_LABELS[latest_year]}').font = TITLE_FONT
ws4.cell(1, 1).alignment = CENTER
ws4.row_dimensions[1].height = 40

# KPI Cards - Row 1
kpi_cards = [
    ('إجمالي الطلاب', totals['students']),
    ('عدد البرامج', totals['programs']),
    ('الطلاب الجدد', totals['new']),
    ('الخريجين', totals['grads']),
]
kpi_cards2 = [
    ('نسبة التخرج في الوقت', pct(totals['grads_ontime'], totals['new_4_ago'])),
    ('نسبة الاستبقاء', pct(totals['retained'], totals['prev_new'])),
    ('نسبة الذكور', pct(totals['male'], totals['students'])),
    ('نسبة الدوليين', pct(totals['intl'], totals['students'])),
]

# Row 3: KPI Labels
# Row 4: KPI Values
ws4.row_dimensions[3].height = 25
ws4.row_dimensions[4].height = 35

for i, (label, val) in enumerate(kpi_cards):
    col = i * 2 + 1
    ws4.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
    ws4.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
    cell_label = ws4.cell(3, col, label)
    cell_label.font = Font(name='Tajawal', bold=True, size=11, color='FFFFFF')
    cell_label.fill = HEADER_FILL
    cell_label.alignment = CENTER
    cell_label.border = THICK_BORDER
    cell_val = ws4.cell(4, col, val)
    cell_val.font = Font(name='Tajawal', bold=True, size=18, color='0D6E6E')
    cell_val.alignment = CENTER
    cell_val.border = THICK_BORDER
    cell_val.fill = LIGHT_FILL
    cell_val.number_format = '#,##0'

# Row 6-7: percentage KPIs
ws4.row_dimensions[6].height = 25
ws4.row_dimensions[7].height = 35

for i, (label, val) in enumerate(kpi_cards2):
    col = i * 2 + 1
    ws4.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+1)
    ws4.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col+1)
    cell_label = ws4.cell(6, col, label)
    cell_label.font = Font(name='Tajawal', bold=True, size=11, color='FFFFFF')
    cell_label.fill = PatternFill('solid', fgColor='C9A227')
    cell_label.alignment = CENTER
    cell_label.border = THICK_BORDER
    cell_val = ws4.cell(7, col, val if val else 'غ/م')
    if val is not None:
        cell_val.number_format = '0.0%'
        if val >= 0.7:
            cell_val.fill = GREEN_FILL
        elif val >= 0.4:
            cell_val.fill = YELLOW_FILL
        else:
            cell_val.fill = RED_FILL
    cell_val.font = Font(name='Tajawal', bold=True, size=18, color='333333')
    cell_val.alignment = CENTER
    cell_val.border = THICK_BORDER

# ── Section 2: Program Details Table for Latest Year ──
ws4.merge_cells('A9:H9')
ws4.cell(9, 1, f'تفاصيل البرامج - {YEAR_LABELS[latest_year]}').font = SUBTITLE_FONT
ws4.cell(9, 1).alignment = CENTER
ws4.row_dimensions[9].height = 30

detail_headers = [
    'البرنامج', 'الدرجة', 'الطلاب', 'جدد', 'مستمرون',
    'خريجون', 'نسبة التخرج', 'نسبة الاستبقاء'
]
for i, h in enumerate(detail_headers, 1):
    ws4.cell(10, i, h)
style_header_row(ws4, 10, len(detail_headers))

row_n = 11
for idx, d in enumerate(sorted(latest_data, key=lambda x: -x['students_total'])):
    is_alt = idx % 2 == 1
    grad_r = pct(d['graduates_ontime'], d['new_4_ago_count'])
    ret_r = pct(d['students_retained'], d['prev_new_count'])

    vals = [
        d['major'], d['degree'],
        d['students_total'], d['students_new'], d['students_retained'],
        d['graduates_total'], grad_r, ret_r
    ]
    for c, v in enumerate(vals, 1):
        ws4.cell(row_n, c, v if v is not None else 'غ/م')
        is_pct = c >= 7
        is_num = 3 <= c <= 6
        style_data_cell(ws4, row_n, c, is_alt, is_num, is_pct)

        # Color-code rates
        if is_pct and v is not None and isinstance(v, float):
            cell = ws4.cell(row_n, c)
            if v >= 0.7:
                cell.fill = GREEN_FILL
            elif v >= 0.4:
                cell.fill = YELLOW_FILL
            else:
                cell.fill = RED_FILL
    row_n += 1

# ── Section 3: Year-over-Year Comparison ──
comp_start = row_n + 2
ws4.merge_cells(f'A{comp_start}:H{comp_start}')
ws4.cell(comp_start, 1, 'مقارنة سنوية - إجمالي الكلية').font = SUBTITLE_FONT
ws4.cell(comp_start, 1).alignment = CENTER
ws4.row_dimensions[comp_start].height = 30

comp_headers = [
    'السنة', 'البرامج', 'الطلاب', 'جدد', 'خريجون',
    'نسبة التخرج', 'نسبة الاستبقاء', 'التغير %'
]
for i, h in enumerate(comp_headers, 1):
    ws4.cell(comp_start + 1, i, h)
style_header_row(ws4, comp_start + 1, len(comp_headers))

row_n = comp_start + 2
prev_total = None
for y in YEAR_ORDER:
    if y not in year_agg:
        continue
    a = year_agg[y]
    is_alt = (row_n - comp_start - 2) % 2 == 1

    grad_r = pct(a['grads_ontime'], a['new_4_ago'])
    ret_r = pct(a['retained'], a['prev_new'])
    change = pct(a['total'] - prev_total, prev_total) if prev_total else None

    vals = [
        YEAR_LABELS[y], a['programs'], a['total'], a['new'], a['grads'],
        grad_r, ret_r, change
    ]
    for c, v in enumerate(vals, 1):
        ws4.cell(row_n, c, v if v is not None else 'غ/م')
        is_pct = c >= 6
        is_num = 2 <= c <= 5
        style_data_cell(ws4, row_n, c, is_alt, is_num, is_pct)
    prev_total = a['total']
    row_n += 1

# ── Section 4: Degree-level breakdown for latest year ──
deg_start = row_n + 2
ws4.merge_cells(f'A{deg_start}:F{deg_start}')
ws4.cell(deg_start, 1, f'توزيع حسب الدرجة العلمية - {YEAR_LABELS[latest_year]}').font = SUBTITLE_FONT
ws4.cell(deg_start, 1).alignment = CENTER
ws4.row_dimensions[deg_start].height = 30

deg_headers = ['الدرجة', 'عدد البرامج', 'الطلاب', 'الخريجين', 'نسبة التخرج', 'نسبة الاستبقاء']
for i, h in enumerate(deg_headers, 1):
    ws4.cell(deg_start + 1, i, h)
style_header_row(ws4, deg_start + 1, len(deg_headers))

degree_agg = {}
for d in latest_data:
    deg = d['degree']
    if deg not in degree_agg:
        degree_agg[deg] = {
            'count': 0, 'total': 0, 'grads': 0,
            'grads_ontime': 0, 'new_4_ago': 0,
            'retained': 0, 'prev_new': 0
        }
    da = degree_agg[deg]
    da['count'] += 1
    da['total'] += d['students_total']
    da['grads'] += d['graduates_total']
    da['grads_ontime'] += d['graduates_ontime']
    da['new_4_ago'] += d['new_4_ago_count']
    da['retained'] += d['students_retained']
    da['prev_new'] += d['prev_new_count']

row_n = deg_start + 2
deg_order = ['بكالوريوس', 'الماجستير', 'دكتوراه']
for idx, deg in enumerate(deg_order):
    if deg not in degree_agg:
        continue
    da = degree_agg[deg]
    is_alt = idx % 2 == 1
    grad_r = pct(da['grads_ontime'], da['new_4_ago'])
    ret_r = pct(da['retained'], da['prev_new'])

    vals = [deg, da['count'], da['total'], da['grads'], grad_r, ret_r]
    for c, v in enumerate(vals, 1):
        ws4.cell(row_n, c, v if v is not None else 'غ/م')
        is_pct = c >= 5
        is_num = 2 <= c <= 4
        style_data_cell(ws4, row_n, c, is_alt, is_num, is_pct)
    row_n += 1

# ── Section 5: Gender distribution for latest year ──
gen_start = row_n + 2
ws4.merge_cells(f'A{gen_start}:D{gen_start}')
ws4.cell(gen_start, 1, f'التوزيع حسب الجنس - {YEAR_LABELS[latest_year]}').font = SUBTITLE_FONT
ws4.cell(gen_start, 1).alignment = CENTER

gen_headers = ['الفئة', 'العدد', 'النسبة', '']
for i, h in enumerate(gen_headers, 1):
    ws4.cell(gen_start + 1, i, h)
style_header_row(ws4, gen_start + 1, 3)

gen_data = [
    ('ذكور', totals['male'], pct(totals['male'], totals['students'])),
    ('إناث', totals['female'], pct(totals['female'], totals['students'])),
    ('سعوديون', totals['saudi'], pct(totals['saudi'], totals['students'])),
    ('دوليون', totals['intl'], pct(totals['intl'], totals['students'])),
]
row_n = gen_start + 2
for idx, (label, val, rate) in enumerate(gen_data):
    is_alt = idx % 2 == 1
    ws4.cell(row_n, 1, label)
    style_data_cell(ws4, row_n, 1, is_alt)
    ws4.cell(row_n, 2, val)
    style_data_cell(ws4, row_n, 2, is_alt, is_num=True)
    ws4.cell(row_n, 3, rate)
    style_data_cell(ws4, row_n, 3, is_alt, is_pct=True)
    row_n += 1

# Set column widths for dashboard
for col in range(1, 9):
    ws4.column_dimensions[get_column_letter(col)].width = 18

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  SHEET 5: تطور البرامج (Program Trends)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws5 = wb.create_sheet('تطور البرامج')
ws5.sheet_view.rightToLeft = True

ws5.merge_cells('A1:L1')
ws5.cell(1, 1, 'تطور أعداد الطلاب حسب البرنامج عبر السنوات').font = TITLE_FONT
ws5.cell(1, 1).alignment = CENTER
ws5.row_dimensions[1].height = 35

# Pivot: programs as rows, years as columns
trend_headers = ['البرنامج', 'الدرجة'] + [YEAR_LABELS[y] for y in YEAR_ORDER]
for i, h in enumerate(trend_headers, 1):
    ws5.cell(2, i, h)
style_header_row(ws5, 2, len(trend_headers))

row_n = 3
for (dept, major, degree), rows in sorted(programs.items()):
    is_alt = (row_n - 3) % 2 == 1
    ws5.cell(row_n, 1, major)
    style_data_cell(ws5, row_n, 1, is_alt)
    ws5.cell(row_n, 2, degree)
    style_data_cell(ws5, row_n, 2, is_alt)

    year_map = {d['sem']: d['students_total'] for d in rows}
    for ci, y in enumerate(YEAR_ORDER):
        val = year_map.get(y, '')
        ws5.cell(row_n, ci + 3, val if val != '' else '')
        style_data_cell(ws5, row_n, ci + 3, is_alt, is_num=True)
    row_n += 1

auto_width(ws5)

# Add trend chart
trend_chart = LineChart()
trend_chart.title = 'تطور أعداد الطلاب'
trend_chart.y_axis.title = 'عدد الطلاب'
trend_chart.style = 10
trend_chart.width = 25
trend_chart.height = 15

# Only plot bachelor programs (they have large numbers)
bach_rows = []
r = 3
for (dept, major, degree), rows in sorted(programs.items()):
    if degree == 'بكالوريوس':
        bach_rows.append(r)
    r += 1

cats = Reference(ws5, min_col=3, max_col=len(YEAR_ORDER)+2, min_row=2)
for br in bach_rows:
    values = Reference(ws5, min_col=3, max_col=len(YEAR_ORDER)+2, min_row=br)
    trend_chart.add_data(values, from_rows=True)
    series = trend_chart.series[-1]
    series.tx = SeriesLabel(v=ws5.cell(br, 1).value)
trend_chart.set_categories(Reference(ws5, min_col=3, max_col=len(YEAR_ORDER)+2, min_row=2))
ws5.add_chart(trend_chart, f'A{row_n + 2}')

print(f'Sheet 5: {row_n - 3} program trends')

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  SHEET 6: حسابات المؤشرات (KPI Calculations)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws6 = wb.create_sheet('حسابات المؤشرات')
ws6.sheet_view.rightToLeft = True

ws6.merge_cells('A1:M1')
ws6.cell(1, 1, 'حسابات المؤشرات التفصيلية لجميع البرامج والسنوات').font = TITLE_FONT
ws6.cell(1, 1).alignment = CENTER
ws6.row_dimensions[1].height = 35

calc_headers = [
    'البرنامج', 'الدرجة', 'السنة',
    'إجمالي الطلاب', 'طلاب جدد', 'مستمرون',
    'خريجون', 'خريجون في الوقت',
    'جدد السنة السابقة', 'حجم دفعة التخرج',
    'نسبة التخرج في الوقت', 'نسبة الاستبقاء',
    'معادلة التخرج', 'معادلة الاستبقاء'
]
# Actually keep it at 12 cols, add formula explanation column
calc_headers2 = [
    'البرنامج', 'الدرجة', 'السنة',
    'إجمالي', 'جدد', 'مستمرون',
    'خريجون', 'في الوقت',
    'جدد سابق', 'حجم الدفعة',
    'نسبة التخرج', 'نسبة الاستبقاء'
]
for i, h in enumerate(calc_headers2, 1):
    ws6.cell(2, i, h)
style_header_row(ws6, 2, len(calc_headers2))

row_n = 3
display_years = [y for y in YEAR_ORDER if y != 38]
for (dept, major, degree), rows in sorted(programs.items()):
    for d in sorted(rows, key=lambda x: x['sem']):
        if d['sem'] == 38:
            continue
        is_alt = (row_n - 3) % 2 == 1

        grad_r = pct(d['graduates_ontime'], d['new_4_ago_count'])
        ret_r = pct(d['students_retained'], d['prev_new_count'])

        vals = [
            d['major'], d['degree'], YEAR_LABELS.get(d['sem'], str(1400+d['sem'])),
            d['students_total'], d['students_new'], d['students_retained'],
            d['graduates_total'], d['graduates_ontime'],
            d['prev_new_count'], d['new_4_ago_count'],
            grad_r, ret_r
        ]
        for c, v in enumerate(vals, 1):
            ws6.cell(row_n, c, v if v is not None else '')
            is_pct = c >= 11
            is_num = 4 <= c <= 10
            style_data_cell(ws6, row_n, c, is_alt, is_num, is_pct)
        row_n += 1

auto_width(ws6)
ws6.auto_filter.ref = f'A2:L{row_n - 1}'

# Add formula explanation
formula_start = row_n + 2
ws6.merge_cells(f'A{formula_start}:L{formula_start}')
ws6.cell(formula_start, 1, 'شرح المعادلات').font = SUBTITLE_FONT
ws6.cell(formula_start, 1).alignment = CENTER

formulas = [
    ('نسبة التخرج في الوقت المحدد', 'خريجون في الوقت ÷ حجم دفعة التخرج × 100 (بكالوريوس: جدد قبل 4 سنوات، ماجستير: قبل سنتين، دكتوراه: قبل 3 سنوات)'),
    ('نسبة الاستبقاء (السنة الأولى)', 'الطلاب المستمرون ÷ عدد الطلاب الجدد في السنة السابقة × 100'),
    ('الطلاب الجدد', 'الطلاب الموجودون في السنة الحالية وغير موجودين في السنة السابقة'),
    ('الطلاب المستمرون', 'عدد الطلاب الجدد في السنة السابقة الذين استمروا في السنة الحالية'),
    ('حجم دفعة التخرج', 'عدد المستجدين في سنة بداية الدفعة (تختلف حسب الدرجة: بكالوريوس 4 سنوات، ماجستير سنتان، دكتوراه 3 سنوات)'),
    ('تسلسل السنوات', '1438(أساس) → 1439 → 1440 → 1441 → 1442 → 1444 → 1445 → 1446 → 1447'),
    ('ملاحظة', 'السنة 1443 غير موجودة (مدمجة مع 1442) - عند الحساب إذا وقعت السنة المستهدفة على 1443 يتم الرجوع لأقرب سنة متاحة (1442)'),
]
for i, (label, desc) in enumerate(formulas):
    r = formula_start + 1 + i
    ws6.cell(r, 1, label)
    ws6.cell(r, 1).font = Font(name='Tajawal', bold=True, size=11, color='0D6E6E')
    ws6.merge_cells(f'B{r}:L{r}')
    ws6.cell(r, 2, desc)
    ws6.cell(r, 2).font = DATA_FONT
    ws6.cell(r, 2).alignment = RIGHT

print(f'Sheet 6: {row_n - 3} calculation rows')

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  SHEET 7: سجل الخريجين (Graduate Records)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws7 = wb.create_sheet('سجل الخريجين')
ws7.sheet_view.rightToLeft = True

ws7.merge_cells('A1:L1')
ws7.cell(1, 1, 'سجل الخريجين التفصيلي - لمتابعة التواصل والتغذية الراجعة').font = TITLE_FONT
ws7.cell(1, 1).alignment = CENTER
ws7.row_dimensions[1].height = 35

grad_headers = [
    'السنة', 'الرقم الجامعي', 'الاسم', 'التخصص', 'الدرجة', 'القسم',
    'الجنس', 'الجنسية', 'تاريخ القبول', 'تاريخ التخرج',
    'تاريخ التخرج المتوقع', 'المعدل'
]
for i, h in enumerate(grad_headers, 1):
    ws7.cell(2, i, h)
style_header_row(ws7, 2, len(grad_headers))

# Read graduates CSV
grad_rows = []
if os.path.exists(GRADS_CSV):
    with open(GRADS_CSV, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=';')
        for r in reader:
            grad_rows.append(r)

row_n = 3
for idx, g in enumerate(grad_rows):
    is_alt = idx % 2 == 1
    year_val = int(g['السنة']) if g['السنة'] else 0
    year_label = YEAR_LABELS.get(year_val, str(1400 + year_val) if year_val else '')

    vals = [
        year_label,
        g['الرقم_الجامعي'],
        g['الاسم'],
        g['التخصص'],
        g['الدرجة'],
        g['القسم'],
        g['الجنس'],
        g['الجنسية'],
        g['تاريخ_القبول'],
        g['تاريخ_التخرج'],
        g['تاريخ_التخرج_المتوقع'],
        g['المعدل'],
    ]
    for c, v in enumerate(vals, 1):
        ws7.cell(row_n, c, v)
        style_data_cell(ws7, row_n, c, is_alt)
    row_n += 1

auto_width(ws7, min_w=14, max_w=35)
ws7.auto_filter.ref = f'A2:L{row_n - 1}'

# Freeze top rows for easy scrolling
ws7.freeze_panes = 'A3'

# ── Summary section below data ──
summary_start = row_n + 2
ws7.merge_cells(f'A{summary_start}:F{summary_start}')
ws7.cell(summary_start, 1, 'ملخص الخريجين حسب السنة والدرجة').font = SUBTITLE_FONT
ws7.cell(summary_start, 1).alignment = CENTER

# Aggregate graduates by year and degree
grad_summary = {}
for g in grad_rows:
    year_val = int(g['السنة']) if g['السنة'] else 0
    degree = g['الدرجة']
    key = (year_val, degree)
    grad_summary[key] = grad_summary.get(key, 0) + 1

sum_headers = ['السنة', 'الدرجة', 'عدد الخريجين']
for i, h in enumerate(sum_headers, 1):
    ws7.cell(summary_start + 1, i, h)
style_header_row(ws7, summary_start + 1, len(sum_headers))

sr = summary_start + 2
for (y, deg) in sorted(grad_summary.keys()):
    is_alt = (sr - summary_start - 2) % 2 == 1
    ws7.cell(sr, 1, YEAR_LABELS.get(y, str(1400 + y)))
    style_data_cell(ws7, sr, 1, is_alt)
    ws7.cell(sr, 2, deg)
    style_data_cell(ws7, sr, 2, is_alt)
    ws7.cell(sr, 3, grad_summary[(y, deg)])
    style_data_cell(ws7, sr, 3, is_alt, is_num=True)
    sr += 1

print(f'Sheet 7: {len(grad_rows)} graduate records')

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  SHEET 8: غير المكملين (Non-Completers)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws8 = wb.create_sheet('غير المكملين')
ws8.sheet_view.rightToLeft = True

ws8.merge_cells('A1:L1')
ws8.cell(1, 1, 'سجل الطلاب غير المكملين - لدراسة أحوالهم ومتابعة حالاتهم').font = TITLE_FONT
ws8.cell(1, 1).alignment = CENTER
ws8.row_dimensions[1].height = 35

# Status color coding
STATUS_COLORS = {
    'منسحب': PatternFill('solid', fgColor='FFF3CD'),       # أصفر فاتح
    'مؤجل': PatternFill('solid', fgColor='D1ECF1'),        # أزرق فاتح
    'مؤجل قبول': PatternFill('solid', fgColor='D1ECF1'),   # أزرق فاتح
    'معتذر': PatternFill('solid', fgColor='E2E3E5'),       # رمادي فاتح
    'منقطع عن الدراسة': PatternFill('solid', fgColor='F8D7DA'),  # أحمر فاتح
    'مفصول اكاديميا': PatternFill('solid', fgColor='F5C6CB'),    # أحمر
    'مطوي قيده': PatternFill('solid', fgColor='F8D7DA'),         # أحمر فاتح
    'موقوف تأديبي / مف': PatternFill('solid', fgColor='F5C6CB'),  # أحمر
    'متوفى': PatternFill('solid', fgColor='D6D8DB'),             # رمادي
}

nc_headers = [
    'آخر سنة ظهور', 'الرقم الجامعي', 'الاسم', 'التخصص', 'الدرجة', 'القسم',
    'الحالة', 'الجنس', 'الجنسية', 'تاريخ القبول', 'المعدل', 'نوع الدراسة'
]
for i, h in enumerate(nc_headers, 1):
    ws8.cell(2, i, h)
style_header_row(ws8, 2, len(nc_headers))

# Read non-completers CSV
nc_rows = []
if os.path.exists(NONCOMP_CSV):
    with open(NONCOMP_CSV, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=';')
        for r in reader:
            nc_rows.append(r)

row_n = 3
for idx, nc in enumerate(nc_rows):
    is_alt = idx % 2 == 1
    year_val = int(nc['آخر_سنة']) if nc['آخر_سنة'] else 0
    year_label = YEAR_LABELS.get(year_val, str(1400 + year_val) if year_val else '')

    vals = [
        year_label,
        nc['الرقم_الجامعي'],
        nc['الاسم'],
        nc['التخصص'],
        nc['الدرجة'],
        nc['القسم'],
        nc['الحالة'],
        nc['الجنس'],
        nc['الجنسية'],
        nc['تاريخ_القبول'],
        nc['المعدل'],
        nc['نوع_الدراسة'],
    ]
    for c, v in enumerate(vals, 1):
        ws8.cell(row_n, c, v)
        style_data_cell(ws8, row_n, c, is_alt)

    # تلوين خلية الحالة حسب نوعها
    status_val = nc['الحالة']
    if status_val in STATUS_COLORS:
        ws8.cell(row_n, 7).fill = STATUS_COLORS[status_val]

    row_n += 1

auto_width(ws8, min_w=14, max_w=35)
ws8.auto_filter.ref = f'A2:L{row_n - 1}'

# Freeze top rows
ws8.freeze_panes = 'A3'

# ── Summary: count by status ──
nc_summary_start = row_n + 2
ws8.merge_cells(f'A{nc_summary_start}:D{nc_summary_start}')
ws8.cell(nc_summary_start, 1, 'توزيع غير المكملين حسب الحالة').font = SUBTITLE_FONT
ws8.cell(nc_summary_start, 1).alignment = CENTER

status_count = {}
for nc in nc_rows:
    st = nc['الحالة']
    status_count[st] = status_count.get(st, 0) + 1

sc_headers = ['الحالة', 'العدد', 'النسبة']
for i, h in enumerate(sc_headers, 1):
    ws8.cell(nc_summary_start + 1, i, h)
style_header_row(ws8, nc_summary_start + 1, len(sc_headers))

total_nc = len(nc_rows)
sr = nc_summary_start + 2
for st in sorted(status_count.keys(), key=lambda x: -status_count[x]):
    is_alt = (sr - nc_summary_start - 2) % 2 == 1
    ws8.cell(sr, 1, st)
    style_data_cell(ws8, sr, 1, is_alt)
    if st in STATUS_COLORS:
        ws8.cell(sr, 1).fill = STATUS_COLORS[st]
    ws8.cell(sr, 2, status_count[st])
    style_data_cell(ws8, sr, 2, is_alt, is_num=True)
    ws8.cell(sr, 3, pct(status_count[st], total_nc))
    style_data_cell(ws8, sr, 3, is_alt, is_pct=True)
    sr += 1

# Total row
ws8.cell(sr, 1, 'الإجمالي')
ws8.cell(sr, 1).font = Font(name='Tajawal', bold=True, size=11)
ws8.cell(sr, 1).alignment = CENTER
ws8.cell(sr, 1).border = THICK_BORDER
ws8.cell(sr, 2, total_nc)
ws8.cell(sr, 2).font = Font(name='Tajawal', bold=True, size=11)
ws8.cell(sr, 2).alignment = CENTER
ws8.cell(sr, 2).border = THICK_BORDER
ws8.cell(sr, 2).number_format = '#,##0'

# ── Summary: count by status and year ──
nc_year_start = sr + 3
ws8.merge_cells(f'A{nc_year_start}:F{nc_year_start}')
ws8.cell(nc_year_start, 1, 'توزيع غير المكملين حسب السنة والحالة').font = SUBTITLE_FONT
ws8.cell(nc_year_start, 1).alignment = CENTER

year_status_count = {}
for nc in nc_rows:
    year_val = int(nc['آخر_سنة']) if nc['آخر_سنة'] else 0
    st = nc['الحالة']
    key = (year_val, st)
    year_status_count[key] = year_status_count.get(key, 0) + 1

ys_headers = ['السنة', 'الحالة', 'العدد']
for i, h in enumerate(ys_headers, 1):
    ws8.cell(nc_year_start + 1, i, h)
style_header_row(ws8, nc_year_start + 1, len(ys_headers))

sr = nc_year_start + 2
for (y, st) in sorted(year_status_count.keys()):
    is_alt = (sr - nc_year_start - 2) % 2 == 1
    ws8.cell(sr, 1, YEAR_LABELS.get(y, str(1400 + y)))
    style_data_cell(ws8, sr, 1, is_alt)
    ws8.cell(sr, 2, st)
    style_data_cell(ws8, sr, 2, is_alt)
    if st in STATUS_COLORS:
        ws8.cell(sr, 2).fill = STATUS_COLORS[st]
    ws8.cell(sr, 3, year_status_count[(y, st)])
    style_data_cell(ws8, sr, 3, is_alt, is_num=True)
    sr += 1

# ── Summary: count by program ──
nc_prog_start = sr + 3
ws8.merge_cells(f'A{nc_prog_start}:F{nc_prog_start}')
ws8.cell(nc_prog_start, 1, 'توزيع غير المكملين حسب البرنامج').font = SUBTITLE_FONT
ws8.cell(nc_prog_start, 1).alignment = CENTER

prog_count = {}
for nc in nc_rows:
    key = (nc['التخصص'], nc['الدرجة'])
    prog_count[key] = prog_count.get(key, 0) + 1

pp_headers = ['التخصص', 'الدرجة', 'العدد', 'النسبة']
for i, h in enumerate(pp_headers, 1):
    ws8.cell(nc_prog_start + 1, i, h)
style_header_row(ws8, nc_prog_start + 1, len(pp_headers))

sr = nc_prog_start + 2
for (prog, deg) in sorted(prog_count.keys(), key=lambda x: -prog_count[x]):
    is_alt = (sr - nc_prog_start - 2) % 2 == 1
    ws8.cell(sr, 1, prog)
    style_data_cell(ws8, sr, 1, is_alt)
    ws8.cell(sr, 2, deg)
    style_data_cell(ws8, sr, 2, is_alt)
    ws8.cell(sr, 3, prog_count[(prog, deg)])
    style_data_cell(ws8, sr, 3, is_alt, is_num=True)
    ws8.cell(sr, 4, pct(prog_count[(prog, deg)], total_nc))
    style_data_cell(ws8, sr, 4, is_alt, is_pct=True)
    sr += 1

print(f'Sheet 8: {len(nc_rows)} non-completer records')

# ── Move Dashboard sheet to first position ──
wb.move_sheet('لوحة المعلومات', offset=-5)

# ── Save ──
wb.save(OUT_PATH)
print(f'\nSaved: {OUT_PATH}')
print('Done!')
